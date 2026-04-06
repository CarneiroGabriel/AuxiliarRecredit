from __future__ import annotations

import io
import re
import urllib.parse
from typing import List

import pandas as pd
import streamlit as st

from importador_core_final_v5 import (
    DEFAULT_CONFIG,
    ImportadorRecredit,
    aplicar_regra_unidade_txt,
    carregar_config_json,
    aplicar_vencimento_txt,
)

st.set_page_config(page_title="Recredit - Importador + Cobrança", page_icon="🏢", layout="wide")


# =========================
# Helpers importador
# =========================
def boletos_para_df(boletos):
    rows = []
    for b in boletos:
        rows.append(
            {
                "unidade": b.unidade,
                "vencimento": b.vencimento,
                "qtde_rateios": len(b.rateios),
                "historico": b.historico,
                "valor_total": round(b.total, 2),
                "rateios": " | ".join(
                    f"{r.codigo} - {r.historico}: {r.valor:.2f}".replace(".", ",")
                    for r in b.rateios
                ),
            }
        )
    return pd.DataFrame(rows)


def txt_para_download_bytes(texto: str, encoding: str = "cp1252") -> bytes:
    try:
        return texto.encode(encoding)
    except Exception:
        return texto.encode("utf-8")


# =========================
# Helpers cobrança whatsapp
# =========================
def get_pdf_reader():
    erros = []
    try:
        from pypdf import PdfReader
        return PdfReader
    except Exception as e:
        erros.append(f"pypdf: {e}")
    try:
        import pdfplumber

        class _PdfPlumberReader:
            def __init__(self, fileobj):
                self._pdf = pdfplumber.open(fileobj)
                self.pages = self._pdf.pages

        return _PdfPlumberReader
    except Exception as e:
        erros.append(f"pdfplumber: {e}")
    raise RuntimeError("Não foi possível carregar uma biblioteca de leitura de PDF. " + " | ".join(erros))


def br_to_float(value: str) -> float:
    return float(value.replace(".", "").replace(",", "."))


def fmt_currency(value: float | None) -> str:
    if value is None:
        return ""
    s = f"{value:,.2f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")


def clean_name(name: str) -> str:
    name = re.sub(r"\*.*?\*", "", name).strip()
    name = re.sub(r"\s+", " ", name)
    return name


def first_name(full_name: str) -> str:
    cleaned = clean_name(full_name)
    parts = cleaned.split()
    return parts[0].title() if parts else ""


def limpar_texto(texto: str) -> str:
    texto = texto.replace("\xa0", " ")
    texto = texto.replace("\r", "\n")
    texto = re.sub(r"[ \t]+", " ", texto)
    return texto


def extrair_condominio(texto: str) -> str:
    padrao = re.compile(r"Condom[ií]nio:\s*\(\d+\)\s*-\s*(.+)", flags=re.IGNORECASE)
    for linha in texto.splitlines():
        linha = linha.strip()
        m = padrao.search(linha)
        if m:
            return m.group(1).strip()
    return "Seu condomínio"


def parse_condomino_line(line: str) -> tuple[str, str, str]:
    content = line.replace("Condômino Atual:", "").strip()
    phone = ""
    document = ""

    phone_match = re.search(r"\s*-\s*Fone:(.*)$", content)
    if phone_match:
        phone = phone_match.group(1).strip()
        content = content[: phone_match.start()].rstrip()

    doc_match = re.search(r"\s*-\s*(CPF|CNPJ):\s*([0-9./-]+)\s*$", content)
    if doc_match:
        document = doc_match.group(2).strip()
        content = content[: doc_match.start()].rstrip()

    name = content.strip()
    return name, document, phone


def extract_phone_candidates(phone_raw: str) -> List[str]:
    if not phone_raw:
        return []

    chunks = re.split(r"[;/]| e ", phone_raw, flags=re.IGNORECASE)
    phones: List[str] = []

    for chunk in chunks:
        digits = re.sub(r"\D", "", chunk)

        if len(digits) >= 13 and digits.startswith("55"):
            digits = digits[2:]

        if len(digits) > 11:
            tail_match = re.search(r"(\d{2}9?\d{8})$", digits)
            if tail_match:
                digits = tail_match.group(1)

        if len(digits) in (10, 11):
            phones.append(digits)

    deduped: List[str] = []
    for phone in phones:
        if phone not in deduped:
            deduped.append(phone)

    return deduped


def parse_total_value(total_line: str) -> float | None:
    if not total_line:
        return None

    match = re.search(r"(-?[\d.]+,\d{2})\d*$", total_line)
    if match:
        return br_to_float(match.group(1))
    return None


def classify_charge(entries: list[dict]) -> str:
    statuses = {entry["status"] for entry in entries if entry["status"]}

    if "E" in statuses:
        return "execucao"
    if "J" in statuses:
        return "juridica"
    if "S" in statuses or "A" in statuses:
        return "acordo"
    return "amistosa"


def normalize_unit(unit_raw: str) -> str:
    unit = re.sub(r"\s+", " ", str(unit_raw).strip().upper())
    unit = unit.replace("CONDOMÍONIO", "CONDOMINIO")
    unit = unit.replace("CONDOMÍNIO", "CONDOMINIO")
    unit = unit.replace("CONDIMÍNIO", "CONDOMINIO")
    unit = unit.replace("SL. ", "SL ")
    unit = unit.replace("SALA ", "SL ")
    unit = re.sub(r"\s+", " ", unit).strip()

    # mantém códigos colados e com sufixos, ex.: 00103, 101T2, 405T2, 0101C
    if re.fullmatch(r"\d{5}", unit):
        return unit
    if re.fullmatch(r"\d{3,4}[A-Z]{0,2}\d{0,2}", unit):
        return unit

    # 101 01 -> 01-101
    m = re.fullmatch(r"(\d{3,4})\s+(\d{2})", unit)
    if m:
        return f"{m.group(2)}-{m.group(1)}"

    # BL 1 102 -> BL 1 102
    m = re.fullmatch(r"BL\s*(\d+)\s+(\d{3,4}[A-Z]{0,2}\d{0,2})", unit)
    if m:
        return f"BL {m.group(1)} {m.group(2)}"

    # sala variations
    m = re.fullmatch(r"([A-Z0-9.-]+)\s+SL\s*(\d{1,3})", unit)
    if m:
        return f"{m.group(1)}-SL{int(m.group(2)):02d}"
    m = re.fullmatch(r"SL\s*(\d{1,3})", unit)
    if m:
        return f"SL {int(m.group(1)):02d}"

    return unit


UNIT_PATTERN = r"(?:BL\s*\d+\s+\d{3,4}[A-Z]{0,2}\d{0,2}|B\.[A-Z]-\d{3}|B\d+-SL\d{2}|B\d+-\d{3}|[A-Z]-\d{2,3}|\d{5}|\d{3,4}[A-Z]{0,2}\d{0,2}|\d{2}-\d{3}|\d{3,4}\s+\d{2}|SL\.?\s*\d{1,3}|SALA\s*\d{1,3})"
ENTRY_RE = re.compile(
    rf"^(?:(?P<status>[A-Z])\s+)?(?P<unit>{UNIT_PATTERN})\s+"
    r"(?P<desc>.+?)\s+(?P<due>\d{2}/\d{2}/\d{4})\s+[-\d., ]+$",
    flags=re.IGNORECASE,
)


def parse_pdf_bytes_cobranca(
    pdf_bytes: bytes,
    msg_amistosa: str,
    msg_acordo: str,
    msg_juridica: str,
    msg_execucao: str,
) -> pd.DataFrame:
    PdfReader = get_pdf_reader()
    reader = PdfReader(io.BytesIO(pdf_bytes))
    full_text = []
    raw_lines = []

    for page in reader.pages:
        page_text = page.extract_text() or ""
        full_text.append(page_text)
        raw_lines.extend([line.strip() for line in page_text.splitlines() if line.strip()])

    texto = limpar_texto("\n".join(full_text))
    condominio = extrair_condominio(texto)

    ignored_prefixes = (
        "Unidade",
        "Sistema de Condomínios",
        "Atualização de Boletos",
        "Página ",
        "Histórico do Lançamento",
        "Condomínio:",
        "Da Unidade",
        "Da emissão",
        "até ",
        "KGR RECEBIMENTOS",
        "RECREDIT - RECEBIMENTOS",
        "Total:",
        "J. Em cobrança judicial",
        "CONDOMINIO EDIFICIO JOSE PHILIPPS",
        "Demonstrativo",
        "Período de ",
        "( * ) SALDO ANTERIOR",
        "( + ) RECEITAS",
        "( - ) DESPESAS",
        "( = ) SALDO ATUAL",
        "* Informamos que",
        "Para serviços on-line",
        "Autenticação Mecânica",
        "Nome do Beneficiário",
        "Local de Pagamento",
        "Data do Documento",
        "Nosso Número",
        "Uso do Banco",
        "Instruções:",
        "Sacador/Avalista",
    )

    lines = [
        line
        for line in raw_lines
        if line
        and not any(line.startswith(prefix) for prefix in ignored_prefixes)
        and "<PARSED TEXT FOR PAGE:" not in line
    ]

    rows = []
    current_entries = []

    for i, line in enumerate(lines):
        entry_match = ENTRY_RE.match(line)
        if entry_match:
            current_entries.append(
                {
                    "status": (entry_match.group("status") or "").upper(),
                    "unit": normalize_unit(entry_match.group("unit")),
                    "desc": entry_match.group("desc").strip(),
                    "due": entry_match.group("due"),
                }
            )
            continue

        if line == "* Total Unidade *":
            if not current_entries:
                continue

            cond_line = lines[i + 1] if i + 1 < len(lines) else ""
            address_line = lines[i + 2] if i + 2 < len(lines) else ""
            total_line = lines[i + 3] if i + 3 < len(lines) else ""

            # alguns arquivos trazem a linha-resumo na mesma linha do total unidade ou pulam endereço/condômino
            if not cond_line.startswith("Condômino Atual:"):
                # procura próximas linhas relevantes
                for j in range(i + 1, min(i + 6, len(lines))):
                    if lines[j].startswith("Condômino Atual:"):
                        cond_line = lines[j]
                        address_line = lines[j + 1] if j + 1 < len(lines) else ""
                        total_line = lines[j + 2] if j + 2 < len(lines) else ""
                        break

            unit = current_entries[-1]["unit"]
            name_raw, document, phone_raw = parse_condomino_line(cond_line)
            name = clean_name(name_raw)
            f_name = first_name(name_raw)
            phones = extract_phone_candidates(phone_raw)
            phone_main = phones[0] if phones else ""
            charge_type = classify_charge(current_entries)
            total_value = parse_total_value(total_line)
            due_dates = [entry["due"] for entry in current_entries]
            message = build_message(
                charge_type,
                f_name,
                unit,
                condominio,
                due_dates,
                total_value,
                msg_amistosa,
                msg_acordo,
                msg_juridica,
                msg_execucao,
            )

            if charge_type == "execucao" or not phone_main:
                whatsapp_link = ""
            else:
                whatsapp_link = f"https://wa.me/55{phone_main}?text={urllib.parse.quote(message)}"

            rows.append(
                {
                    "condominio": condominio,
                    "unidade": unit,
                    "nome": name,
                    "primeiro_nome": f_name,
                    "documento": document,
                    "telefone_bruto": phone_raw,
                    "telefone_cadastro": phone_main,
                    "telefones_encontrados": " | ".join(phones),
                    "tipo_cobranca": charge_type,
                    "qtde_titulos": len(current_entries),
                    "vencimentos_aberto": " | ".join(due_dates),
                    "resumo_lancamentos": " | ".join(entry["desc"] for entry in current_entries),
                    "valor_total_atualizado": total_value,
                    "endereco": address_line.replace("Endereço:", "").strip(),
                    "link_whatsapp": whatsapp_link,
                    "mensagem_whatsapp": message,
                }
            )

            current_entries = []

    # fallback específico boleto único como jose philipps
    if not rows:
        unit_m = re.search(r"\((\d{3,5}[A-Z-]*)\)\s+([A-ZÁÉÍÓÚÂÊÔÃÕÇ' ]+)\s*-\s*\d{11,14}", texto, flags=re.IGNORECASE)
        due_m = re.search(r"Data de Vencimento\s+.*?\n.*?(\d{2}/\d{2}/\d{4})", texto, flags=re.DOTALL)
        total_m = re.search(r"\(=\) Valor do Documento\s+.*?\n.*?(\d+[\.,]\d{2})", texto, flags=re.DOTALL)
        phone_m = re.search(r"Usuário:.*?\n.*?\n.*?Nome do Pagador.*?\n.*?\((\d{3,5}[A-Z-]*)\)\s+([^\n]+)", texto, flags=re.DOTALL)
        cond_name = re.search(r"CONDOMINIO .*?\n", texto)
        if unit_m and due_m and total_m:
            unit = normalize_unit(unit_m.group(1))
            name = clean_name(unit_m.group(2))
            due = due_m.group(1)
            total_value = br_to_float(total_m.group(1).replace('.', '').replace(',', '.')) if ',' in total_m.group(1) else None
            message = build_message(
                "amistosa", first_name(name), unit, condominio, [due], total_value,
                msg_amistosa, msg_acordo, msg_juridica, msg_execucao,
            )
            rows.append(
                {
                    "condominio": condominio,
                    "unidade": unit,
                    "nome": name,
                    "primeiro_nome": first_name(name),
                    "documento": "",
                    "telefone_bruto": "",
                    "telefone_cadastro": "",
                    "telefones_encontrados": "",
                    "tipo_cobranca": "amistosa",
                    "qtde_titulos": 1,
                    "vencimentos_aberto": due,
                    "resumo_lancamentos": "Boleto em aberto",
                    "valor_total_atualizado": total_value,
                    "endereco": "",
                    "link_whatsapp": "",
                    "mensagem_whatsapp": message,
                }
            )

    df = pd.DataFrame(rows)
    if df.empty:
        return df

    order = {"execucao": 0, "juridica": 1, "acordo": 2, "amistosa": 3}
    df["_ordem_tipo"] = df["tipo_cobranca"].map(order).fillna(99)
    df = df.sort_values(["_ordem_tipo", "unidade"]).drop(columns=["_ordem_tipo"])
    return df


def build_message(
    kind: str,
    first: str,
    unit: str,
    condominio: str,
    due_dates: list[str],
    total_value: float | None,
    msg_amistosa: str,
    msg_acordo: str,
    msg_juridica: str,
    msg_execucao: str,
) -> str:
    due_text = ", ".join(due_dates)
    total_text = fmt_currency(total_value)

    placeholders = {
        "{primeiro_nome}": first,
        "{unidade}": unit,
        "{condominio}": condominio,
        "{vencimentos}": due_text,
        "{valor_total}": total_text,
    }

    template_map = {
        "amistosa": msg_amistosa,
        "acordo": msg_acordo,
        "juridica": msg_juridica,
        "execucao": msg_execucao,
    }
    message = template_map.get(kind, msg_amistosa)

    for chave, valor in placeholders.items():
        message = message.replace(chave, valor)
    return message


def df_to_excel_bytes(df: pd.DataFrame) -> bytes:
    from openpyxl.styles import Font

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="dados")
        ws = writer.sheets["dados"]
        headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
        link_col = headers.get("link_whatsapp")
        if link_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=link_col)
                link = cell.value
                if link:
                    cell.hyperlink = link
                    cell.value = "Abrir WhatsApp"
                    cell.style = "Hyperlink"
                    cell.font = Font(color="0563C1", underline="single")
    return output.getvalue()


# =========================
# Interface
# =========================
st.title("🏢 Recredit - Importador e Cobrança")

aba1, aba2, aba3 = st.tabs([
    "Converter PDF/Excel em TXT",
    "Aplicar regra em TXT",
    "Cobrança WhatsApp",
])

with aba1:
    st.subheader("Converter espelho em TXT de importação")
    col1, col2 = st.columns([1, 1])

    with col1:
        arquivo = st.file_uploader(
            "Arquivo de entrada",
            type=["pdf", "xlsx", "xlsm", "xls"],
            key="arquivo_importador",
        )
        config_json = st.file_uploader(
            "Configuração JSON opcional",
            type=["json"],
            key="config_importador",
        )
        regra_unidade = st.selectbox(
            "Regra de unidade",
            options=["padrao", "portal", "mafalda", "orlando"],
            index=0,
            key="regra_importador",
        )
        codigo_orlando = st.text_input("Código alvo Orlando", value="101", key="orlando_importador")
        novo_vencimento = st.text_input("Substituir vencimento (opcional)", value="", placeholder="dd/mm/aaaa", key="novo_vencimento_importador")

    with col2:
        st.caption("Cabeçalho do TXT")
        emp_codigo = st.text_input("EMP_CODIGO", value=DEFAULT_CONFIG["emp_codigo"])
        emp_nomemp = st.text_input("EMP_NOMEMP", value=DEFAULT_CONFIG["emp_nomemp"])
        emp_cegece = st.text_input("EMP_CEGECE", value=DEFAULT_CONFIG["emp_cegece"])
        administradora_nome = st.text_input("Administradora nome", value=DEFAULT_CONFIG["administradora_nome"])
        administradora_fantasia = st.text_input("Administradora fantasia", value=DEFAULT_CONFIG["administradora_fantasia"])
        blq_datemi = st.text_input("BLQ_DATEMI", value=DEFAULT_CONFIG["blq_datemi"])

    if arquivo is not None:
        try:
            config_extra = carregar_config_json(config_json.read()) if config_json else {}
            config = {
                **config_extra,
                "emp_codigo": emp_codigo,
                "emp_nomemp": emp_nomemp,
                "emp_cegece": emp_cegece,
                "administradora_nome": administradora_nome,
                "administradora_fantasia": administradora_fantasia,
                "blq_datemi": blq_datemi,
            }
            core = ImportadorRecredit(config=config)
            boletos = core.parse_file(arquivo.read(), arquivo.name)
            boletos = core.aplicar_vencimento(boletos, novo_vencimento)

            if not boletos:
                st.warning("Nenhum boleto reconhecido.")
            else:
                txt_saida = core.gerar_txt(boletos)
                txt_saida = aplicar_regra_unidade_txt(txt_saida, regra_unidade, codigo_orlando)
                df_boletos = boletos_para_df(boletos)
                st.success(f"{len(df_boletos)} boletos reconhecidos.")
                st.dataframe(df_boletos, use_container_width=True)
                st.download_button(
                    "Baixar TXT",
                    data=txt_para_download_bytes(txt_saida),
                    file_name=f"{arquivo.name.rsplit('.', 1)[0]}.txt",
                    mime="text/plain",
                )
        except Exception as e:
            st.error(f"Erro ao processar arquivo: {e}")

with aba2:
    st.subheader("Aplicar regra em TXT já existente")
    txt_file = st.file_uploader("Envie o TXT", type=["txt"], key="arquivo_txt_regra")
    regra_txt = st.selectbox("Regra", options=["padrao", "portal", "mafalda", "orlando"], key="regra_txt")
    codigo_orlando_txt = st.text_input("Código alvo Orlando", value="101", key="orlando_txt")
    novo_vencimento_txt = st.text_input("Substituir vencimento (opcional)", value="", placeholder="dd/mm/aaaa", key="novo_vencimento_txt")

    if txt_file is not None:
        try:
            conteudo = txt_file.read().decode("cp1252", errors="ignore")
            convertido = aplicar_regra_unidade_txt(conteudo, regra_txt, codigo_orlando_txt)
            convertido = aplicar_vencimento_txt(convertido, novo_vencimento_txt)
            st.download_button(
                "Baixar TXT corrigido",
                data=txt_para_download_bytes(convertido),
                file_name=f"corrigido_{txt_file.name}",
                mime="text/plain",
            )
            st.text_area("Prévia", convertido[:4000], height=300)
        except Exception as e:
            st.error(f"Erro ao aplicar regra: {e}")

with aba3:
    st.subheader("Cobrança WhatsApp")
    st.caption("Suporta unidades como 0101, 01-101, 101 01, BL 1 102, B1-104, B.A-103, 0101C, A-12, B2-SL04 e similares.")

    col_a, col_b = st.columns([1, 1])
    with col_a:
        pdf_inad = st.file_uploader("PDF de inadimplência", type=["pdf"], key="pdf_inad")

    with col_b:
        msg_amistosa = st.text_area(
            "Mensagem amistosa",
            value="Olá, {primeiro_nome}. Identificamos em {condominio} pendências da unidade {unidade}, com vencimento(s) em {vencimentos}. Podemos te ajudar na regularização.",
            height=110,
        )
        msg_acordo = st.text_area(
            "Mensagem acordo",
            value="Olá, {primeiro_nome}. Consta em {condominio} pendência relacionada ao acordo/unidade {unidade}, com vencimento(s) em {vencimentos}. Pedimos retorno para regularização.",
            height=110,
        )
        msg_juridica = st.text_area(
            "Mensagem jurídica",
            value="Olá, {primeiro_nome}. A unidade {unidade} de {condominio} possui pendência(s) em aberto, com vencimento(s) em {vencimentos}. \n\n Esta unidade possui apontamento em cobrança jurídica. Solicitamos contato para tratativa.",
            height=110,
        )
        msg_execucao = st.text_area(
            "Mensagem execução",
            value="Olá, {primeiro_nome}. A unidade {unidade} de {condominio} possui débito em execução. Vencimento(s): {vencimentos}. Valor atualizado aproximado: R$ {valor_total}.",
            height=110,
        )

    if pdf_inad is not None:
        try:
            df = parse_pdf_bytes_cobranca(pdf_inad.read(), msg_amistosa, msg_acordo, msg_juridica, msg_execucao)
            if df.empty:
                st.warning("Nenhuma unidade reconhecida neste PDF.")
            else:
                st.success(f"{len(df)} unidades reconhecidas.")
                st.dataframe(df, use_container_width=True)
                st.download_button(
                    "Baixar Excel",
                    data=df_to_excel_bytes(df),
                    file_name="cobranca_whatsapp.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
        except Exception as e:
            st.error(f"Erro ao processar PDF de inadimplência: {e}")
